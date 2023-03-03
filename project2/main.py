

from game import Game
from player import Player
from chessDataBase import ChessDataBase
from document import Document
from node import Node
from tree import Tree
from ete3 import Tree as ETE3Tree
from bigtree import list_to_tree, print_tree, print_tree, find, findall, tree_to_dot
from bigtree import Node as N
import re
import matplotlib.pyplot as plt
import copy
import openpyxl




import math

irregularOpenings = {
    "Birds Opening": "f4,d5",
    "English opening": "c4,e5",
    "Dunst Opening|(Scicilian Defence)": "Nc3, c5"
}
#1.e4, 1.d4:
regularOpenings = {
    "Queen's Pawn Game|(Indian Defence)": "d4,Nf6",
    "Queen's Pawn Game|(Closed Game)": "d4,d5",
    "Queen's Pawn Game|()": "",
    "Queen's Pawn Game|()": "",



}

# #Task 2
def importGame(path, number):
    game = None
    try:
        file = open(path, "r")
        #Find line where new game starts
        breakPoints = _getBreakPoints(path)
        lines = file.readlines()
        gameNr = 1

        for i in range(len(lines)):#len(lines)
            if i in breakPoints and gameNr == number:
                metadata = lines[i:i+13]
                # #Move lines from-to
                moveFrom = i+13
                nextStart = breakPoints[gameNr]
                moveTo = nextStart
                moves = lines[moveFrom:moveTo-1]
                #Create game
                game = _createGameFromLoad(metadata[4], metadata[5])
                game.setMetaData(metadata)
                movePairs = _createMovesFromLoad(moves)
                game.setMoves(movePairs)
            if i in breakPoints:
                gameNr+=1
           
        file.close()
    except:
        print("could not read file")
    
    return game

# #Task 3
def saveGame(path, game):
    try:
        file = open(path, "r")
        file.close()
    except:
        print("could not read file")
    try:
        #Open file in append mode
        file = open(path, "a")
        #Add metadata
        metadata = game._getMetaData()
        file.write(metadata)

        #Add moves
        moves = game._getMoves()
        counter = 0
        for line in moves:
            if counter<80:
                file.write(line)
                counter+=1
            else:
                file.write("\n"+line)
                counter = 0
        file.write("\n")
        file.flush()
        file.close()
    except:
        print("could not append to file")
    return 0


# #Task 4
#Load games
def loadGames(path):
    if type(path) == str:
        return _loadGamesFromPath(path)
    elif type(path) == ChessDataBase:
        return _loadGamesFromDataBase(path)
    
def _getNumberOfLines(path):
    file = open(path, "r")
    lines = 0
    for line in file:
        lines+=1
    file.close()
    return lines

def _getBreakPoints(path):
    file = open(path, "r")
    breakPoints = []
    for i, line in enumerate(file):
        if "[Event" in line:
            breakPoints.append(i)
    file.close()
    numLinnes = _getNumberOfLines(path)
    breakPoints.append(numLinnes)
    return breakPoints

def _loadGamesFromPath(path):
    game = None
    metadata = None
    movePairs = []
    moves = ""
    file = None
    breakPoints = []
    games = ChessDataBase()
    try:
        file = open(path, "r")
        #Find line where new game starts
        breakPoints = _getBreakPoints(path)
        lines = file.readlines()
        gameNr = 0


        for i in range(len(lines)):#len(lines)
            if i in breakPoints:
                metadata = lines[i:i+13]
                #Move lines from-to
                moveFrom = i+13
                nextStart = breakPoints[gameNr+1]
                moveTo = nextStart
                moves = lines[moveFrom:moveTo-1]

                #Create game
                game = _createGameFromLoad(metadata[4], metadata[5])
                game.setMetaData(metadata)
                movePairs = _createMovesFromLoad(moves)
                game.setMoves(movePairs)
                games.addGame(game)

                #Restart variables
                metadata = []
                moves = ""
                game = None
                gameNr+=1

        file.close()
    except:
        print("could not read file")
    
    return games

def _createGameFromLoad(p1N, p2N):
    p1Name = p1N[8:-3]
    p1 = Player(p1Name)
    p2Name = p2N[8:-3]
    p2 = Player(p2Name)
    game = Game(p1,p2)
    return game

def _createMovesFromLoad(moves):
    s = ""
    for i in moves:
        s+=i.replace("\n"," ")

    #Remove everything except moves
    splitted = s.split(" ")
    for i in range(len(splitted)):
        if "." in splitted[i]:
            splitted[i] = " "
        elif "{" in splitted[i]:
            splitted[i] = " "
        elif '}' in splitted[i]:
            splitted[i] = " "
    
    #Sanitize
    list = []
    for i in splitted:
        if i == " " or i == "":
            continue
        else:
            list.append(i)

    #Create return list
    #Collect moves in pairs per turn
    returnList = []
    for i in range(len(list)-2):
        if i%2 == 0:
            returnList.append((list[i],list[i+1]))
    if(len(list)%2 == 1):
        returnList.append((list[len(list)-1],"None"))
    else:
        returnList.append((list[len(list)-2],list[len(list)-1]))

    return returnList

def _loadGamesFromDataBase(path):
    games = path.getGames()
    chessGames = []
    for i in games:
        chessGames.append(i)
    return chessGames

#Save games
def saveGames(dataBase, games):
    for game in games:
        dataBase.addGame(game)

def exportGames(path, database):
    for game in database.getGames():
        saveGame(path, game)

# #Task 5
#Load data from Excel file
def loadFromExcel(path):
    print(" ")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    maxCol = ws.max_column
    maxRow = ws.max_row

    gameContainer = [[],[]]
    placer = 0
    for i in range(1,maxRow+1):
        row = ""
        for j in range(2,maxCol+1):
            cell = ws.cell(row=i, column=j)
            row+=str(cell.value)+" "
        #Shift to moves
        if row == "None None ":
            placer = 1
        #Remove header of moves
        elif row == "White Black ":
            continue
        else:
            gameContainer[placer].append(row)
    
    #Create metadata
    metaData = []
    for data in gameContainer[0]:
        metaData.append(data[:-1])

    #Create moves
    moves = []
    for data in gameContainer[1]:
        d = data.split(" ")
        moves.append((d[0],d[1]))

    #Create game
    whitePlayer = Player(metaData[4][6:])
    blackPlayer = Player(metaData[5][6:])
    game = Game(whitePlayer, blackPlayer)
    game.setMetaDataExcel(metaData)
    game.setMoves(moves)
    return game


#Save data to Excel file
def saveToExcel(game):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Chess"
    metaNameColumn = "B"
    metaValueColumn = "C"
    metaData = game.getMetaData()
    if len(metaData)==12:
        metaDataNames = ["Event", "Site", "Date", "Round", "White", "Black", "Result", "ECO", "Opening", "PlyCount", "WhiteElo", "BlackElo"]
    elif len(metaData)==13:
        metaDataNames = ["Event", "Site", "Date", "Round", "White", "Black", "Result", "ECO", "Opening", "Variation", "PlyCount", "WhiteElo", "BlackElo"]

    # print("len(metaDataNames): ", len(metaDataNames))
    # print("len(metaData): ", len(metaData))

    lineCount = 1
    #Add metadata
    for i in range(len(metaDataNames)):
        name = metaDataNames[i]
        value = metaData[i]
        if type(value) == Player:
            value = value.getName()
        metaNamePlacement = metaNameColumn+str(i+1)
        metaValuePlacement = metaValueColumn+str(i+1)
        ws[metaNamePlacement] = name
        ws[metaValuePlacement] = value
        lineCount+=1

    #Add moves
    moves = game.getMoves()
    # print("moves: ", moves)
    turnCount = 1
    lineCount+=1
    turnColumn = "A"
    whiteColumn = "B"
    blackColumn = "C"
    ws[turnColumn+str(lineCount)] = "Turn"
    ws[whiteColumn+str(lineCount)] = "White"
    ws[blackColumn+str(lineCount)] = "Black"
    lineCount+=1
    for i in range(lineCount,lineCount+len(moves)):
        turn = turnCount
        white = moves[turnCount-1][0]
        black = moves[turnCount-1][1]

        ws[turnColumn+str(i)] = turn
        ws[whiteColumn+str(i)] = white
        ws[blackColumn+str(i)] = black
        turnCount+=1

    workbook.save("project2/DataFiles/Chess.xlsx")


# #Task 7
def statisticsStockfish(database):
    games = database.getGames()
    print("Number of games: ", len(games))
    winnerWhite = 0
    drawWhite = 0
    loserWhite = 0
    winnerBlack = 0
    drawBlack = 0
    loserBlack = 0
    totalWinner = 0
    totalDraw = 0
    totalLoser = 0

    for game in games:
        #Get color of Stockfish
        color = None
        if game.getWhite().getName() == "Stockfish 15 64-bit":
            color = "White"
        elif game.getBlack().getName() == "Stockfish 15 64-bit":
            color = "Black"

        # print(color)
        
        #get winner
        result = game.getWinner()[0]
        winner = game.getWinner()[1]
        
        # print(result, winner)
        
        if result == "1-0":
            if color == "White":
                winnerWhite+=1
            elif color == "Black":
                loserBlack+=1
        elif result == "0-1":
            if color == "White":
                loserWhite+=1
            elif color == "Black":
                winnerBlack+=1
        elif result == "1/2-1/2":
            if color == "White":
                drawWhite+=1
            elif color == "Black":
                drawBlack+=1
        


        # if result == "1/2-1/2":
        #     print("")

    totalWinner = winnerWhite+winnerBlack
    totalDraw = drawWhite+drawBlack
    totalLoser = loserWhite+loserBlack

    print("White: ", winnerWhite, drawWhite, loserWhite)
    print("Black: ", winnerBlack, drawBlack, loserBlack)
    print("Total: ", totalWinner, totalDraw, totalLoser)

    return winnerWhite, drawWhite, loserWhite, winnerBlack, drawBlack, loserBlack, totalWinner, totalDraw, totalLoser

# #Task 8
def createPlots(database):
    #All games
    AG = _allGames(copy.copy(database))

    #Stockfish with white
    SW= _stockfishWhite(copy.copy(database))

    #Stockfish with black
    SB = _stockfishBlack(copy.copy(database))

    #Stockfish winning
    SW = _stockfishWinning(copy.copy(database))

    #Stockfish losing
    SL = _stockfishLosing(copy.copy(database))

    return AG, SW, SB, SW, SL
    #All games

def _allGames(database):
    games = database.getGames()
    moves = {}

    for game in games:
        numMoves = game.getNumberOfMoves()
        if numMoves in moves:
            moves[numMoves]+=1
        else:
            moves[numMoves] = 1
    
    return _plotOfNumberOfMoves(games,moves)

def _stockfishWhite(database):
    games = database.getGames()
    moves = {}
    for game in games:
        if game.getWhite().getName() == "Stockfish 15 64-bit":
            numMoves = game.getNumberOfMoves()
            if numMoves in moves:
                moves[numMoves]+=1
            else:
                moves[numMoves] = 1
    return _plotOfNumberOfMoves(games,moves)

def _stockfishBlack(database):
    games = database.getGames()
    moves = {}
    for game in games:
        if game.getBlack().getName() == "Stockfish 15 64-bit":
            numMoves = game.getNumberOfMoves()
            if numMoves in moves:
                moves[numMoves]+=1
            else:
                moves[numMoves] = 1
    return _plotOfNumberOfMoves(games,moves)    

def _stockfishWinning(database):
    games = database.getGames()
    moves = {}
    for game in games:
        if game.getWinner()[1] == "Stockfish 15 64-bit":
            numMoves = game.getNumberOfMoves()
            if numMoves in moves:
                moves[numMoves]+=1
            else:
                moves[numMoves] = 1
    return _plotOfNumberOfMoves(games,moves)

def _stockfishLosing(database):
    games = database.getGames()
    moves = {}
    for game in games:
        if game.getWinner()[1] != "Stockfish 15 64-bit" and game.getWinner()[1] != "Draw":
            numMoves = game.getNumberOfMoves()
            if numMoves in moves:
                moves[numMoves]+=1
            else:
                moves[numMoves] = 1
    return _plotOfNumberOfMoves(games,moves)

def _plotOfNumberOfMoves(games, moves):
    #Assuumptions: one turn = two moves

    #Sort dictionary
    sortedDict = {}
    for i in range(len(moves)):
        minValue = min(moves.keys())
        sortedDict[minValue] = moves[minValue]
        moves.pop(minValue)

    yValues = []
    minValue = min(sortedDict.keys())
    maxValue = max(sortedDict.keys())
    for i in sortedDict.keys():#range(minValue, maxValue+1)
        keys = sortedDict.keys()
        value = 0
        for k in keys:
            if k>=i:
                value+=sortedDict[k]
        yValues.append(value)

    #Mean, standard deviation
    mean = 0
    for i in sortedDict.keys():
        mean+=i*sortedDict[i]
    mean = round(mean/len(sortedDict.values()),2)

    std = 0
    for i in sortedDict.keys():
        std+=(i-mean)**2*sortedDict[i]
    std = round(math.sqrt(std/len(games)),2)

    return sortedDict, yValues, mean, std



# Task 10
def createTreesFromGames(games):
    trees = []
    for game in games:
        # Adding first game to tree
        if len(trees)==0:
            t = Tree(game)
            trees.append(t)
        #Adding rest of games to tree
        else:
            addedNew = False
            for tree in trees:
                if tree.getRoot().getNodeValue() == game.getMoves()[0][0]:
                    tree.addNewGameToTree(game)
                    addedNew = True
                    break
            if addedNew == False:
                t = Tree(game)
                trees.append(t)
    return trees



def _printTree(tree):

    if type(tree)== Tree:
        structure = tree.createTreeStructure()
        root = list_to_tree(structure)
        print_tree(root)
    elif type(tree)==list:
        print("list")
        print(tree)
        # root = list_to_tree(tree)
        # print_tree(root)
    else:
        print("Wrong input")

#Task 11
def createTreesOfDepth(t, depth):
    trees = _getTreesOfDepth(t, depth)
    returnValue = copy.copy(trees)
    winnerStat = []
    for t in trees:
        name = str(t.getRoot().getNodeValue())
        saveTreeAsPng(t, name)
        winners = t.getWinnersFromLeafs()
        winnerStat.append([t.getRoot().getNodeValue(),winners])
        # print("Winners tree: "+str(t.getRoot().getNodeValue()),winners)
    # winnersTable = []
    # createWinnersTable(winnerStat[0])
    return returnValue, winnerStat

def _getTreesOfDepth(t, depth):
    # Assumptions: depth = number of moves
    
    trees = copy.copy(t)
    newTrees = []
    for tree in trees:
        #Remove all nodes that are not in the depth
        newTree = tree.removeNodesUnderGivenDepth(depth)
        newTrees.append(newTree)
    return trees
            
def saveTreeAsPng(tree, name):
    if(type(tree) == Tree):
        structure = tree.createTreeStructure()
        root = list_to_tree(structure)
        graph = tree_to_dot(root)
        graph.write_png('project2/trees/'+name+'.png')
    elif type(tree)==list:
        root = list_to_tree(tree)
        graph = tree_to_dot(root)
        graph.write_png('project2/trees/'+name+'.png') 

def createWinnersTable(winnerStat):
    print(winnerStat)

# Task 11







if __name__ == "__main__":
    path = "project2/datafiles/Stockfish_15_64-bit.commented.2600.pgn"
    path100 = "project2/datafiles/games.txt"
    # #Task 1
    # p1 = Player("Stockfish")
    # p1.setColor("White")
    # p2 = Player("Mathias")
    # p2.setColor("Black")
    # g = Game(p1,p2)
    # g.createBoard()

    # g.movePiece(p1,"Pa2a4")


    # # #Task 2
    # print("------------------")
    # print("Task 2")
    # print("------------------")
    # print("Importing game...")
    # game = importGame(path,3)#Path to games and game number
    # print("Game imported:")
    # print(game)

    # # #Task 3
    # print("------------------")
    # print("Task 3")
    # print("------------------")
    # print("Saving game to file...")
    # saveGame("notes.txt",game)#"project2/datafiles/games.txt"
    # print("Game saved to file")

    # #Task 4
    # print("------------------")
    # print("Task 4")
    # print("------------------")
    # print("Loading games from file...")
    # games = loadGames(path)
    # print("Games loaded from file")
    # print("Number of games loaded: ", len(games.getGames()))

    # print("Saving games to another database...")
    # exportGames("task4.txt",games)
    # print("Games saved.")


    # #Task 5
    # print("------------------")
    # print("Task 5")
    # print("------------------")
    # print("Loading game from file...")
    # game = importGame(path,3)
    # print("Game loaded from file")
    # print("Saving game to excel file...")
    # saveToExcel(game)
    # print("Game saved to excel file")
    # print("Load game from excel file...")
    # game2 = loadFromExcel("project2/datafiles/Chess.xlsx")
    # print(game2)


    # #Task 10
    # print("------------------")
    # print("Task 10")
    # print("------------------")
    # print("Loading games from file...")
    # games = loadGames(path)
    # print("Games loaded from file")
    # print("Number of games loaded: ", len(games.getGames()))
    # print("Creating trees...")
    # trees = createTreesFromGames(games.getGames())
    # print("Trees created")
    
    # #Task 11
    print("------------------")
    print("Task 11")
    print("------------------")


    print("Loading games from file...")
    games = loadGames(path)
    print("Games loaded from file")
    print("Number of games loaded: ", len(games.getGames()))
    print("Creating trees...")
    trees = createTreesFromGames(games.getGames())
    print("Trees created")
    print("Getting trees of depth ...")
    treesDepth, winnersTable = createTreesOfDepth(trees,3)
    # print(len(treesDepth))
    # print(treesDepth[0].createTreeStructure())
    # i = 1

    

    # for t in treesDepth:
    #     name = str(t.getRoot().getNodeValue())
    #     saveTreeAsPng(t, name)
    #     winners = t.getWinnersFromLeafs()
    #     print("Winners tree: "+str(t.getRoot().getNodeValue()),winners)

    #     i+=1
    # i = 1
    # for tree in trees:
    #     print("tree"+str(i)+": ",tree.createTreeStructure())
    #     winners = t.getWinnersFromLeafs(5)
    #     print(winners)
    #     # _printTree(tree)
    #     i+=1


    # #Task 12
    print("------------------")
    print("Task 12")
    print("------------------")
    print("Extracting winners from trees...")
    for t in treesDepth:
        timesPlayed = t.getTimesPlayedLeaf()
        print("Times played tree: "+str(t.getRoot().getNodeValue()),timesPlayed)
    print("Winners extracted")


    # # #Task 7
    # print("------------------")
    # print("Task 7")
    # print("------------------")
    # print("Creating tables over statistics...")
    # database = loadGames(path)
    # stat = statisticsStockfish(database)
    # print("Tables made")
  
    # # #Task 8
    # print("------------------")
    # print("Task 8")
    # print("------------------")
    # print("Creating plots over statistics...")
    # moves = createPlots(database)
    # print("Plots made")

    # #Task 6
    print("------------------")
    print("Task 6")
    print("------------------")
    print("Creating document...")
    doc = Document("Stockfish")
    # doc.createStatTable(stat)
    # doc.createPlot(moves)
    doc.createWinnersTable(winnersTable)
    doc.write()
    print("Document created")

    