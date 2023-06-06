import openpyxl
from task import Task
from pert import PERT
from printer import Printer
import random
import copy

from sklearn import svm
from sklearn.neighbors import KNeighborsClassifier,  KNeighborsRegressor
from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor

class ML():

    def __init__(self):
        self.standardProjects = {}
        self.projects = {}
        self.riskFactors = [0.8,1.0,1.2,1.4]
        self.printer = Printer()


    ###Functions
    ## Load task from excel file
    def loadFile(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        tasks = []
        for i in range(2, ws.max_row+1):
            id = i-1
            type = ws.cell(row=i, column=1).value
            if type == None :
                continue
            
            code = ws.cell(row=i, column=2).value
            description = ws.cell(row=i, column=3).value
            durations = ws.cell(row=i, column=4).value
            task = Task(id, type, code, description, durations)

            tasks.append(task)
        # Add predecessors and successors
        self._addPreAndSuc(tasks, wb) 
        projectName = path.split("/")[-1].split(".")[0]
        self.standardProjects[projectName] = tasks
    
    def _addPreAndSuc(self,tasks, wb):
        ws = wb.active
        # Add predecessors
        for i in range(2, ws.max_row+1):
            type = ws.cell(row=i, column=1).value
            if type == None :
                continue
            for task in tasks:
                code = ws.cell(row=i, column=2).value
                predecessors = []
                successors = []
                if task.code == code:
                    # print("Code: ", code)
                    # Add predecessors
                    predecessorsStr = ws.cell(row=i, column=5).value
                    if predecessorsStr == None:
                        continue
                    else:
                        predecessorsStr = predecessorsStr.replace(" ", "").split(",")
                    for pre in predecessorsStr:
                        for t in tasks:
                            if t.code == pre:
                                predecessors.append(t)
                    task.setPredecessors(predecessors)

        # Add successorts
        for t in tasks:
            # Find predeccsors to task
            for pre in t.predecessors:
                # for each predeccesor, add task as successor
                for task in tasks:
                    if task.code == pre.code:
                        task.successors.append(t)


    ### Task 4
    ##Preprocess data
    ## Create data set for use in machine learning
    def preprocessData(self):
        for project, tasks in self.standardProjects.items():
            self.preprocess(project,tasks)

    def preprocess(self, projectName, tasks):
        standardProjects = []
        tasksCopy = copy.deepcopy(tasks)
        # Create 4 standard project: one for each risk factor
        standardProjects = self._createStandardProjects(tasksCopy, projectName)

        # # Create 1000 projects per risk factor for each project
        projects = self._create1000Projects(standardProjects)

        # Classify projects: Successfull, Acceptable, Failed
        self._classifyProjects(projects)

    def _createStandardProjects(self, tasks, projectName):
        standardProjects = []
        riskFactors = self.riskFactors
        taskList = []
        for i in range(len(riskFactors)):
            taskList.append(copy.deepcopy(tasks))
        for factor in riskFactors:
            standardTasks = taskList.pop()
            for task in standardTasks:
                oldDurations = task.getDurations()
                newDurations = self._calculateNewDuration(oldDurations, factor)
                task.setDurations(newDurations)
                if newDurations != None:
                    task.setDuration(newDurations[1])

            # #Create Diagram
            diagram = PERT(projectName, standardTasks, factor)
            diagram.executeProject()
            standardProjects.append(diagram)
        return standardProjects

    def _calculateNewDuration(self, duration, factor):
        newDuration = []
        if(type(duration) != list):
            return duration
        newExpected = round(duration[1]*factor,3)
        oldMinimum = duration[0]
        oldExpected = duration[1]
        oldMaximum = duration[2]
        if(newExpected<oldMinimum):
            newDuration = [newExpected, oldExpected, oldMaximum]
        elif(newExpected>oldMaximum):
            newDuration = [oldMinimum, oldExpected, newExpected]
        else:
            newDuration = [oldMinimum, newExpected, oldMaximum]
        return newDuration

    def _create1000Projects(self, standardProjects):
        projects = {}
        riskFactors = self.riskFactors
        for r in riskFactors:
            projects[r] = []
        for i in range(len(riskFactors)):
            base = standardProjects[i]
            for j in range(0,1000):
                proj = copy.deepcopy(base)
                projNewDuration = self._randomDuration(proj)

                projNewDuration.executeProject()

                projects[riskFactors[i]].append(projNewDuration)
                self._addProjectToDict(projNewDuration)
        return projects

    def _randomDuration(self, proj):
        tasks = proj.tasks
        newTasks = []
        for task in tasks:
            if task.getDurations() == None:
                newTasks.append(task)
                continue
            else:
                selectedDuration = random.choice(task.getDurations())
                task.setDuration(selectedDuration)
                newTasks.append(task)
        proj.setTasks(newTasks)
        return proj

    def _addProjectToDict(self, project):
        if project.getName() in self.projects:
            self.projects[project.name].append(project)
        else:
            self.projects[project.name] = [project]
    
    def _classifyProjects(self, projects):
        classification = {"Successfull": [], "Acceptable": [], "Failed": []}
        #Find expected time for projects
        expectedTime = projects[1.0][0].getExpectedTime()
        for key, value in projects.items():
            for p in value:
                finishTime = p.getFinishTimes()[1]
                factor = finishTime/expectedTime
                if factor <1.05:
                    p.projectClass = "Successfull"
                    classification["Successfull"].append(p)
                elif factor <1.15:
                    p.projectClass = "Acceptable"
                    classification["Acceptable"].append(p)
                else:
                    p.projectClass = "Failed"
                    classification["Failed"].append(p)
        # for key, value in classification.items():
        #     print(key, len(value))#, value)


    ## Add gate to project
    def addIntermediateGates(self, projectName, gate):
        for project in self.projects[projectName]:
            project.setIntermediateGate(gate)

    ## Create dict with all project.
    ## For each project, sub project are divided in a 80/20 split for training and test data
    def createInstancesDataClassification(self):
        # Copy data
        data = copy.deepcopy(self.projects)
        # Get attibutes and values for each project
        # Keep factors divided inside project
        instances = self._getAttributesAndValues(data)
        # ##Split into Learning data and Test data
        # #Find size of test data
        size = 0
        for key, value in instances.items():
            for factor, projects in value.items():
                size +=len(instances[key][factor])
        
        #Split learning and test data 80/20 split
        size = size/len(instances.keys())
        learningTestData = self._splitData(instances, size)
        return learningTestData
    
    def _getAttributesAndValues(self, data):
        instances = {}
        for key, factors in data.items():
            if key not in instances:
                instances[key] = {} 
            for factor in factors:
                attributes = self._getAttributes(factor)
                valueClassification, valueRegression = self._getValues(factor)
                factor = factor.riskFactor
                if factor not in instances[key]:
                    instances[key][factor] = []
                instances[key][factor].append([attributes, valueClassification, valueRegression])
        return instances
    
    ## Get attributes for project
    def _getAttributes(self, project):
        intGate = project.getTaskByCode(project.intermediateGate)
        headers, rows = project.getRows()
        completion = rows[4]
        i = 0
        stack = []
        visited = []
        for suc in intGate.successors:
            stack.append(suc)
        
        while stack:
            node = stack.pop(0)
            if node in visited:
                continue
            for suc in node.successors:
                if suc not in visited:
                    stack.append(suc)
            
            visited.append(node)

        
        # Get late end for all tasks
        lateEndDates = []
        for task in visited:
            for i in range(len(headers)):
                if task.code == headers[i]:
                    lateEndDates.append(completion[i])
                    break

        return lateEndDates

    def _getValues(self, project):
        valueClassification = project.getProjectClass()
        valueRegression = project.getFinishTime()
        return valueClassification, valueRegression
        

    def _splitData(self, instances, size):
        learningTestData = {}
        for key, value in instances.items():
            learningTestData[key] = {"LearningData": [], "TestData": []}
        for key, value in instances.items():
            for factor, projects in value.items():
                learningData = []
                testData = []
                while len(projects) > 0:
                    random.shuffle(projects)
                    project = projects.pop()
                    learningSize = len(learningData)
                    if learningSize < (size/4)*0.8 :
                        learningData.append(project)
                    else:
                        testData.append(project)
                for data in learningData:
                    learningTestData[key]["LearningData"].append(data)
                for data in testData:
                    learningTestData[key]["TestData"].append(data)
        return learningTestData


    ### Task 5
    ## Perform classification on dataset
    def classification(self, learningTestData):
        classification = {}
        models = {}
        # Create empty dictionaries for further use
        for project, value in learningTestData.items():
            classification[project] = [],[]
            models[project] = None
        ## Sort out data for each project
        for project, value in learningTestData.items():
            learningData = value["LearningData"]
            for data in learningData:
                classification[project][0].append(data[0])
                classification[project][1].append(data[1])

        ## Train every model
        # 1. Support vector machine
        # 2. K nearest neighbour KNeighborsClassifier
        # 3. Decision tree DecisionTreeClassifier

        models = {}
        for project, value in classification.items():
            models[project] = self.trainModelsC(value)

        ## Test every model
        testResult = {}
        for project, model in models.items():
            res = self.testModelsC(model, learningTestData[project]["TestData"])
            testResult[project] = res
        return testResult
    
    # Train every model with training data
    def trainModelsC(self, value):
        trainingModels = {"SVM": svm.SVC(), "KNeighborsClassifier": KNeighborsClassifier(), "DecisionTreeClassifier": DecisionTreeClassifier()}
        attributes = value[0]
        values = value[1]

        trainedModels = {}
        for name, model in trainingModels.items():
            trainedModel = model.fit(attributes, values)
            trainedModels[name] = trainedModel
        return trainedModels
             
    # Test every model with test data
    def testModelsC(self, models, data):
        testRestult = {}
        for modelName, model in models.items():
            testRes = self._performTestC(model, data)
            testRestult[modelName] = testRes
        return testRestult

    def _performTestC(self, model, testData):
        correct = 0
        wrong = 0
        for data in testData:
            attributes = data[0]
            value = data[1]
            modelRes = model.predict([attributes])
            if modelRes == value:
                correct +=1
            else:
                wrong +=1
        return [correct, wrong] 
    

    ### Task 6
    ##Perform regression on dataset
    def regression(self, learningTestData):
        regression = {}
        models = {}
        # Create empty dictionaries for further use
        for project, value in learningTestData.items():
            regression[project] = [],[]
            models[project] = None
        ## Sort out data for each project
        for project, value in learningTestData.items():
            learningData = value["LearningData"]

            for data in learningData:
                regression[project][0].append(data[0])
                regression[project][1].append(data[2])
        ## Train every model
        # 1. Support vector machine
        # 2. K nearest neighbour KNeighborsRegressor
        # 3. Decision tree DecisionTreeRegressor
        models = {}
        for project, value in regression.items():
            models[project] = self.trainModelsR(value)

        testResult = {}
        for project, model in models.items():

            res = self.testModelsR(model, learningTestData[project]["TestData"])
            testResult[project] = res

        return testResult
    
    # Train every model with training data
    def trainModelsR(self, value):
        trainingModels = {"SVM": svm.SVR(), "KNeighborsRegressor": KNeighborsRegressor(), "DecisionTreeRegressor": DecisionTreeRegressor()}
        attributes = value[0]
        values = value[1]
        trainedModels = {}
        for name, model in trainingModels.items():
            trainedModel = model.fit(attributes, values)
            trainedModels[name] = trainedModel
        return trainedModels

    # Test every model with test data
    def testModelsR(self, models, data):
        testRestult = {}
        for modelName, model in models.items():
            testRes = self._performTestR(model, data)
            testRestult[modelName] = testRes
        return testRestult

    def _performTestR(self, model, testData):
        res = {"ModelRes": [], "Value": [], "Loss": []}
        for data in testData:
            attributes = data[0]
            value = data[2]
            modelRes = model.predict([attributes])[0]
            
            loss = abs(modelRes - value)
            res["ModelRes"].append(modelRes)
            res["Value"].append(value)
            res["Loss"].append(loss)

        return res
        # return [modelRes, value, loss] 
    

    ### For printing
    def printClassification(self, res):
        self.printer.printClassification(res)
    
    def printRegression(self, res):
        self.printer.printRegression(res)