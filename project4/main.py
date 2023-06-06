import openpyxl
from task import Task
from pert import PERT
from machineLearning import ML

# # Task 2
## Load tasks from excel file
def load(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    tasks = []
    for i in range(2, ws.max_row+1):
        id = i-1
        type = ws.cell(row=i, column=1).value
        if type == None :
            continue
        
        code = ws.cell(row=i, column=2).value
        description = ws.cell(row=i, column=3).value
        durations = ws.cell(row=i, column=4).value
        task = Task(id, type, code, description, durations)

        tasks.append(task)
    # Add predecessors and successors
    _addPreAndSuc(tasks, wb)        
    return tasks


def _addPreAndSuc(tasks, wb):
    ws = wb.active
    # Add predecessors
    for i in range(2, ws.max_row+1):
        type = ws.cell(row=i, column=1).value
        if type == None :
            continue
        for task in tasks:
            code = ws.cell(row=i, column=2).value
            predecessors = []
            if task.code == code:
                # print("Code: ", code)
                # Add predecessors
                predecessorsStr = ws.cell(row=i, column=5).value
                if predecessorsStr == None:
                    continue
                else:
                    predecessorsStr = predecessorsStr.replace(" ", "").split(",")
                for pre in predecessorsStr:
                    for t in tasks:
                        if t.code == pre:
                            predecessors.append(t)
                task.setPredecessors(predecessors)

    # Add successorts
    for t in tasks:
        # Find predeccsors to task
        for pre in t.predecessors:
            # for each predeccesor, add task as successor
            for task in tasks:
                if task.code == pre.code:
                    task.successors.append(t)



if  __name__ == "__main__":
    Path_V = "project4/dataFiles/Villa.xlsx"
    Path_W = "project4/dataFiles/Warehouse.xlsx"
    Tasks_V = load(Path_V)
    Tasks_W = load(Path_W)


    ###Task 1
    print("--------------------Task 3--------------------")
    ##Create diagram
    print("Creating test project...\n")
    diagram = PERT("Test Project", Tasks_W)
    ##Execute project
    print("Executing test project...\n")
    diagram.executeProject()

    ##print early dates of p
    print("earlyDates of test project", diagram.earlyDates)
    print("lateDates of thest project", diagram.lateDates)

    ## Print the process plan
    print("Process plan for test project...\n")
    diagram.printProcessPlan()
    
    ## Print early and late dates table
    print("Printing early and late dates for test project...\n")
    diagram.printEarlyAndLateDates()


    #### Machine learning
    ## Create machine learning object
    print("Creating machine learning object...")
    ml = ML()
    print("Machine learning object created\n")

    ## Load data
    print("Loading data...")
    ml.loadFile(Path_V)
    ml.loadFile(Path_W)
    print("Data loaded\n")

    ### Task 4
    print("--------------------Task 4--------------------")
    ## Preprocess data
    print("Preprocessing data...")
    ml.preprocessData()
    print("Data preprocessed\n")


    ## Add gates to each project
    print("Adding Intermediate Gate to each project...")
    ml.addIntermediateGates("Warehouse", "J")
    ml.addIntermediateGates("Villa", "N.1")
    print("Intermediate Gates added\n")

    ## Create set of data for Classification and Regression
    print("Creating set of data for Classification and Regression...")
    learningTestData = ml.createInstancesDataClassification()
    print("Set of data created\n")
    
    ### Task 5
    print("--------------------Task 5--------------------")
    ## Perform classification
    print("Performing classification...")
    res = ml.classification(learningTestData)
    ml.printClassification(res)
    print("Classification performed\n")

    ### Task 6
    print("--------------------Task 6--------------------")
    ## Perform regression
    print("Performing regression...")
    resReg = ml.regression(learningTestData)
    ml.printRegression(resReg)
    print("Regression performed\n")

